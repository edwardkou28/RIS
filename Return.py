#CODE FOR RETURN

import PySimpleGUI as sg
import pandas as pd
import cv2
import pyzbar.pyzbar as zbar
import time
import datetime
import openpyxl
import subprocess
from openpyxl import load_workbook, Workbook

sg.popup('Please scan your Employee ID', title = "Return")
cap = cv2.VideoCapture(0)
font = cv2.FONT_HERSHEY_SIMPLEX 
stopped = False
l=0


while(True):
    ret = cv2.waitKey(1) & 0xFF
    # Capture frame-by-frame
    ret, frame = cap.read()
    decodedObjects = zbar.decode(frame)
    if len(decodedObjects) > 0:
        stopped = True
        for obj in decodedObjects:
            Employee_ID = obj.data.decode("utf-8")
            cv2.putText(frame,Employee_ID, (50, 100), font, 2, (0, 0, 255), 3)
            l+=1

# Display the resulting frame
    cv2.imshow('Scan Employee ID',frame)
    if l == 1:
        cv2.waitKey(1)
        time.sleep(2)
        cv2.destroyAllWindows()
        break 
sg.popup('Please scan your Item ID', title = "Return")
cap = cv2.VideoCapture(0)
font = cv2.FONT_HERSHEY_SIMPLEX 
stopped = False
l=0

while(True):
    ret = cv2.waitKey(1) & 0xFF
    # Capture frame-by-frame
    ret, frame = cap.read()
    decodedObjects = zbar.decode(frame)
    if len(decodedObjects) > 0:
        stopped = True
        for obj in decodedObjects:
            Return = obj.data.decode("utf-8")
            cv2.putText(frame,Return, (50, 100), font, 2, (0, 0, 255), 3)
            l+=1

# Display the resulting frame
    cv2.imshow('Scan Item ID',frame)
    if l == 1:
        cv2.waitKey(1)
        time.sleep(2)
        cv2.destroyAllWindows()
        break

wb=openpyxl.load_workbook(filename = r"C:\PROJECT\RIS\RIS\RIS.xlsx")
ws = wb.active
search_item = Return

for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if search_item == ws.cell(i,j).value:
                #print("Found the value")
                location=(ws.cell(i,j+1).value)
                if location:
                # Define Excel File Path
                    RIS = r"C:\PROJECT\RIS\RIS\RIS.xlsx"
                    df_RIS = pd.read_excel(RIS)
                    time = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
                    df_RIS.loc[df_RIS.SERIAL_NUMBER == Return, ['DATE_RETURN', 'STATUS_RETURN']] = time, Employee_ID
                    df_RIS.to_excel(RIS, index=False)
                    location2 = (ws.cell(i,j+8).value)
                
                    sg.popup("Item has been returned to", location2, "successfully!", title = "OK !")
                    
                else:
                    sg.popup('Error! Item has not been borrowed yet.', title = "Error !")


