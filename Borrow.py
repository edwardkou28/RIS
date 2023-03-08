
#CODE FOR BORROW

import PySimpleGUI as sg
import pandas as pd
import cv2
import pyzbar.pyzbar as zbar
import time
import datetime
import openpyxl
import subprocess
from openpyxl import load_workbook, Workbook

sg.popup('Please scan your Employee ID', title = "Borrow")
cap = cv2.VideoCapture(0)
font = cv2.FONT_HERSHEY_SIMPLEX 
stopped = False
l=0

while(True):

    key = cv2.waitKey(1)
    if key == 27:
        break

    ret = cv2.waitKey(1) & 0xFF
    # Capture frame-by-frame
    ret, frame = cap.read()
    decodedObjects = zbar.decode(frame)
    if len(decodedObjects) > 0:
        stopped = True
        for obj in decodedObjects:
            Employee_ID = obj.data.decode("utf-8")
            cv2.putText(frame,Employee_ID, (50, 100), font, 2, (0, 0, 255), 3)
            l+=1

# Display the resulting frame
    cv2.imshow('Scan Employee ID',frame)
    if l == 1:
        cv2.waitKey(1)
        time.sleep(2)
        cv2.destroyAllWindows()
        break 
sg.popup('Please scan your Item ID', title = "Borrow")
cap = cv2.VideoCapture(0)
font = cv2.FONT_HERSHEY_SIMPLEX 
stopped = False
l=0

while(True):
    ret = cv2.waitKey(1) & 0xFF
    # Capture frame-by-frame
    ret, frame = cap.read()
    decodedObjects = zbar.decode(frame)
    if len(decodedObjects) > 0:
        stopped = True
        for obj in decodedObjects:
            Borrow = obj.data.decode("utf-8")
            cv2.putText(frame,Borrow, (50, 100), font, 2, (0, 0, 255), 3)
            l+=1

# Display the resulting frame
    cv2.imshow('Scan Item ID',frame)
    if l == 1:
        cv2.waitKey(1)
        time.sleep(2)
        cv2.destroyAllWindows()
        break 

wb=openpyxl.load_workbook(filename = r"C:\PROJECT\RIS\RIS\RIS.xlsx")
ws = wb.active
search_item = Borrow

for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if search_item == ws.cell(i,j).value:
                location=(ws.cell(i,j+3).value)
                if location:
                # Define Excel File Path
                    RIS = r"C:\PROJECT\RIS\RIS\RIS.xlsx"
                    df_RIS = pd.read_excel(RIS)
                    time = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
                    df_RIS.loc[df_RIS.SERIAL_NUMBER == Borrow, ['DATE_BORROW','STATUS_BORROW','DATE_RETURN','STATUS_RETURN']] = time, Employee_ID, "", ""
                    df_RIS.to_excel(RIS, index=False)
                    sg.popup('Borrowed item has been recorded successfully!', title = "OK !")

                else:
                    location3 = (ws.cell(i,j+2).value)
                    sg.popup('The above item has been borrowed by', location3 , title = "Error !")
                    
