#CODE FOR REPORT

import PySimpleGUI as sg
import pandas as pd
import cv2
import pyzbar.pyzbar as zbar
import time
import datetime
import openpyxl
import subprocess
from openpyxl import load_workbook, Workbook

sg.popup('Please scan your Employee ID', title = "Report")
cap = cv2.VideoCapture(0)
font = cv2.FONT_HERSHEY_SIMPLEX 
stopped = False
l=0

while(True):
    ret = cv2.waitKey(1) & 0xFF
    # Capture frame-by-frame
    ret, frame = cap.read()
    decodedObjects = zbar.decode(frame)
    if len(decodedObjects) > 0:
        stopped = True
        for obj in decodedObjects:
            Employee_ID = obj.data.decode("utf-8")
            cv2.putText(frame,Employee_ID, (50, 100), font, 2, (0, 0, 255), 3)
            l+=1

# Display the resulting frame
    cv2.imshow('Scan Employee ID',frame)
    if l == 1:
        cv2.waitKey(1)
        time.sleep(2)
        cv2.destroyAllWindows()
        break 

sg.popup('Please scan your Item ID', title = "Report")
cap = cv2.VideoCapture(0)
font = cv2.FONT_HERSHEY_SIMPLEX 
stopped = False
l=0

while(True):
    ret = cv2.waitKey(1) & 0xFF
    # Capture frame-by-frame
    ret, frame = cap.read()
    decodedObjects = zbar.decode(frame)
    if len(decodedObjects) > 0:
        stopped = True
        for obj in decodedObjects:
            Report = obj.data.decode("utf-8")
            cv2.putText(frame,Report, (50, 100), font, 2, (0, 0, 255), 3)
            l+=1

# Display the resulting frame
    cv2.imshow('Scan Item ID',frame)
    if l == 1:
        cv2.waitKey(1)
        time.sleep(2)
        cv2.destroyAllWindows()
        break

wb=openpyxl.load_workbook(filename = r"C:\PROJECT\RIS\RIS\RIS.xlsx")
ws = wb.active
search_item = Report

for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if search_item == ws.cell(i,j).value:
                location=(ws.cell(i,j+5).value)
                if location:
                    sg.popup('Item already has an existing report. Clear report to continue.', title = "Error!")
                else:
                    window2 = sg.Window("Lodge Report", [[sg.Text("Please key in your report")], 
                                                [sg.Input(key='Remarks', enable_events=True)],
                                                [sg.Submit('Submit')]])

                    while True:             
                        event, values = window2.read()
                        if event == sg.WIN_CLOSED:
                            break
                        if event == 'Submit':
                            # Define Excel File Path
                            RIS = r"C:\PROJECT\RIS\RIS\RIS.xlsx"
                            df_RIS = pd.read_excel(RIS)
                            time = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
                            Remarks = str(values['Remarks'])
                            df_RIS.loc[df_RIS.SERIAL_NUMBER == Report, ['ID_REPORTER', 'DATE_REPORT', 'REPORT']] = Employee_ID, time, Remarks
                            df_RIS.to_excel(RIS, index=False)
                            window2.close()
                            sg.popup('Report has been recorded successfully!', title = "OK !")
                            